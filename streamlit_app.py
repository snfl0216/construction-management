import streamlit as st
from sqlalchemy import create_engine, text
import pandas as pd
import altair as alt
from datetime import datetime, date
import io
import calendar as pycal

st.set_page_config(page_title="현장 미수관리 시스템", layout="wide")
st.title("현장 미수관리 시스템")

# ⚠️ 관리자 비밀번호
ADMIN_PASSWORD = "chdan1576**"

CLAIM_TYPES = ["선급금", "기성금", "중도금", "잔금", "추가금", "정산금", "AS", "시공부자재"]

engine = create_engine("sqlite:////tmp/construction_v6.db")
PK = "INTEGER PRIMARY KEY AUTOINCREMENT"

# --------------------------------------------------------------------------
# DB 초기화 — 두 데이터 파이프라인 완전히 분리 (이것도 앱 켜져있는 동안 딱 한 번만 실행)
# --------------------------------------------------------------------------
with engine.connect() as conn:
    # ===== 일일수금관리(이력) 기준 : 기성청구현황 / 캘린더 / 리스크현장 =====
    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS claims (
            id {PK},
            site_name TEXT,
            company_name TEXT,
            manager TEXT,
            claim_type TEXT DEFAULT '기성금',
            claim_date TEXT,
            original_due_date TEXT,
            current_due_date TEXT,
            claim_amount BIGINT DEFAULT 0,
            status TEXT DEFAULT '입금대기',
            last_flagged_due_date TEXT,
            last_remark TEXT
        );
    """))
    conn.commit()
    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS payments (
            id {PK},
            claim_id INTEGER,
            payment_date TEXT,
            payment_amount BIGINT DEFAULT 0
        );
    """))
    conn.commit()
    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS claim_delay_history (
            id {PK},
            claim_id INTEGER,
            event_type TEXT,
            changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            old_due_date TEXT,
            new_due_date TEXT,
            payment_date TEXT,
            delay_days INTEGER DEFAULT 0,
            reason TEXT
        );
    """))
    conn.commit()
    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS claim_checkpoints (
            id {PK},
            claim_id INTEGER,
            checkpoint_date TEXT,
            remark TEXT,
            unpaid_balance BIGINT
        );
    """))
    conn.commit()
    # ===== 현장별 미수관리(미수내역) 기준 : 현장별 미수현황 / 완불현장 / 계약현황 =====
    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS site_receivables (
            id {PK},
            no_number INTEGER,
            division TEXT,
            site_name TEXT,
            company_name TEXT,
            manager TEXT,
            branch TEXT,
            contract_code TEXT,
            contract_date TEXT,
            start_date TEXT,
            completion_date TEXT,
            contract_yearmonth TEXT,
            contract_amount BIGINT DEFAULT 0,
            change_amount BIGINT DEFAULT 0,
            total_paid BIGINT DEFAULT 0,
            unpaid_balance BIGINT DEFAULT 0,
            progress_rate REAL DEFAULT 0,
            invoice_progress_rate REAL DEFAULT 0,
            invoice_issue_rate REAL DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            status_label TEXT
        );
    """))
    conn.commit()
    for col_def in ["no_number INTEGER", "division TEXT"]:
        try:
            conn.execute(text(f"ALTER TABLE site_receivables ADD COLUMN {col_def};"))
            conn.commit()
        except Exception:
            conn.rollback()
    try:
        conn.execute(text("ALTER TABLE claims ADD COLUMN last_remark TEXT;"))
        conn.commit()
    except Exception:
        conn.rollback()
    try:
        conn.execute(text("ALTER TABLE claims ADD COLUMN company_name TEXT;"))
        conn.commit()
    except Exception:
        conn.rollback()
    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS site_receivable_details (
            id {PK},
            site_receivable_id INTEGER,
            detail_type TEXT,
            detail_date TEXT,
            amount BIGINT DEFAULT 0,
            note TEXT
        );
    """))
    conn.commit()
    conn.execute(text(f"""
        CREATE TABLE IF NOT EXISTS contract_status_raw (
            id {PK},
            year INTEGER,
            label TEXT,
            m1 REAL, m2 REAL, m3 REAL, m4 REAL, m5 REAL, m6 REAL,
            m7 REAL, m8 REAL, m9 REAL, m10 REAL, m11 REAL, m12 REAL,
            total REAL
        );
    """))
    conn.commit()

# --------------------------------------------------------------------------
# 공용 함수
# --------------------------------------------------------------------------
def parse_amount(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        if pd.isna(v):
            return 0
        return int(v)
    s = str(v).strip().replace(",", "").replace("원", "").replace(" ", "")
    if s in ("", "-", "nan"):
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def safe_date(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "" or v.lower() == "nat":
            return None
        try:
            return datetime.strptime(v[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    try:
        d = pd.to_datetime(v)
        if pd.isna(d):
            return None
        return d.date()
    except Exception:
        return None


def fmt_money(x):
    try:
        if pd.isna(x):
            return "0"
    except (TypeError, ValueError):
        pass
    try:
        return f"{int(x):,}"
    except (TypeError, ValueError):
        return str(x)


def fmt_eok(won):
    """1억원 단위로 소수점 둘째자리까지 (예: 8,309,000,000 -> '83.09억')"""
    try:
        return f"{won / 1e8:,.2f}억"
    except (TypeError, ValueError):
        return "0억"


def render_metric_cards(items):
    """회색 둥근 카드 여러 개를 한 줄로 그린다. items: [(아이콘, 라벨, 값), ...]"""
    html = "<div style='display:grid;grid-template-columns:repeat(%d,minmax(0,1fr));gap:12px;margin:8px 0 20px 0;'>" % len(items)
    for icon, label, value in items:
        html += (
            "<div style='background:#F1EFE8;border-radius:16px;padding:16px;'>"
            f"<div style='font-size:13px;color:#5F5E5A;margin-bottom:8px;'>{icon} {label}</div>"
            f"<div style='font-size:22px;font-weight:700;color:#2C2C2A;'>{value}</div>"
            "</div>"
        )
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)


PROGRESS_MILESTONES = [(20, "자재발주"), (40, "창호발주"), (60, "창호시공"), (80, "유리시공"), (90, "코킹완료"), (100, "방충망")]


def render_progress_milestones(pct):
    """공정율을 퍼센트 숫자 대신, 자재발주20/창호발주40/창호시공60/유리시공80/코킹완료90/방충망100
    단계가 표시된 막대로 보여준다."""
    pct = max(0, min(100, pct))
    current_label = "시작 전"
    for m_pct, m_label in PROGRESS_MILESTONES:
        if pct >= m_pct:
            current_label = m_label

    ticks_html = ""
    for m_pct, m_label in PROGRESS_MILESTONES:
        is_current = (m_label == current_label)
        color = "var(--text-accent, #185FA5)" if is_current else "#888780"
        weight = "700" if is_current else "400"
        if m_pct <= 20:
            transform = "translateX(0)"
        elif m_pct >= 100:
            transform = "translateX(-100%)"
        else:
            transform = "translateX(-50%)"
        ticks_html += (
            f"<div style='position:absolute;left:{m_pct}%;top:0;width:1px;height:8px;"
            f"background:{'var(--text-accent, #185FA5)' if is_current else '#ccc'};'></div>"
            f"<div style='position:absolute;left:{m_pct}%;top:12px;transform:{transform};"
            f"font-size:11px;color:{color};font-weight:{weight};white-space:nowrap;'>{m_label} {m_pct}%</div>"
        )

    html = f"""
    <div style='background:var(--surface-2);border-radius:16px;border:0.5px solid var(--border);padding:20px;margin-bottom:20px;'>
        <div style='font-size:14px;font-weight:700;margin-bottom:6px;'>공정율 진행 단계</div>
        <div style='font-size:12px;color:var(--text-secondary);margin-bottom:28px;'>현재 {pct}% · {current_label} 단계</div>
        <div style='position:relative;height:10px;background:var(--surface-1);border-radius:999px;margin:0 4px 8px 4px;'>
            <div style='position:absolute;left:0;top:0;height:10px;width:{pct}%;background:var(--fill-accent, #378ADD);border-radius:999px;'></div>
            <div style='position:absolute;left:{pct}%;top:-5px;width:20px;height:20px;margin-left:-10px;
                        border-radius:50%;background:var(--surface-2);border:3px solid var(--fill-accent, #378ADD);'></div>
        </div>
        <div style='position:relative;height:34px;margin:0 4px;'>{ticks_html}</div>
    </div>
    """
    st.markdown(html, unsafe_allow_html=True)


def to_won_from_thousands(v):
    """미수내역 엑셀은 금액이 '천원' 단위로 저장돼있어 *1000 해야 하는데,
    int(v)*1000처럼 먼저 정수로 자르고 곱하면 소수점(600원 단위 등)이 통째로 날아간다.
    반드시 원래 값에 1000을 곱한 다음 반올림해야 한다."""
    if v is None:
        return 0
    try:
        if pd.isna(v):
            return 0
    except (TypeError, ValueError):
        pass
    try:
        return round(float(v) * 1000)
    except (TypeError, ValueError):
        return 0


def fmt_money_cols(df, cols):
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = df[c].apply(fmt_money)
    return df


def render_html_table(df, money_cols=None, left_cols=None, fixed_layout=False, narrow_cols=None, col_max_width=None, wrap_cols=None):
    """헤더는 항상 가운데 정렬(공통), 데이터는 금액=오른쪽/현장명=왼쪽/나머지=가운데.
    td/th에 직접 text-align을 걸면 스트림릿 내부 표 스타일이랑 충돌해서 안 먹는 경우가 있어,
    셀 안에 div를 하나 더 넣어 그 div에서 정렬한다 (표 스타일과 완전히 분리됨).
    fixed_layout=True 이면 컬럼 너비를 균등 고정폭으로 맞춘다 (계약현황처럼 셀 내용이 다 짧을 때만 사용).
    narrow_cols: fixed_layout=True일 때, 지정한 컬럼만 폭을 좁게 잡고 나머진 균등분할 (계약현황의 '현장수' 계열용).
    col_max_width: {컬럼명: 'Npx'} 특정 컬럼만 최대 너비를 제한 (fixed_layout=False인 표에서도 사용 가능).
    wrap_cols: 지정한 컬럼은 줄임표(...) 안 쓰고 텍스트 전체를 줄바꿈해서 다 보여준다."""
    money_cols = set(money_cols or [])
    left_cols = set(left_cols if left_cols is not None else ["현장명"])
    narrow_cols = set(narrow_cols or [])
    col_max_width = col_max_width or {}
    wrap_cols = set(wrap_cols or [])
    d = df.copy()

    def fmt_plain(v):
        if pd.isna(v):
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)

    layout_style = "table-layout:fixed;" if fixed_layout else ""
    min_width_style = f"min-width:{max(80, len(d.columns) * 80)}px;" if fixed_layout else ""
    colgroup = ""
    if fixed_layout and narrow_cols:
        n_narrow = sum(1 for c in d.columns if c in narrow_cols)
        n_normal = len(d.columns) - n_narrow
        narrow_w = 6
        normal_w = (100 - narrow_w * n_narrow) / n_normal if n_normal else 0
        colgroup = "<colgroup>" + "".join(
            f"<col style='width:{narrow_w if c in narrow_cols else normal_w}%'>" for c in d.columns
        ) + "</colgroup>"

    html = f"<div style='overflow-x:auto;'><table style='width:100%;{layout_style}{min_width_style}border-collapse:collapse;font-size:13px;'>{colgroup}"
    html += "<thead><tr>"
    for col in d.columns:
        mw = f"max-width:{col_max_width[col]};" if col in col_max_width else ""
        html += (f"<th style='padding:0;border-bottom:2px solid #ddd;background:#fafafa;'>"
                  f"<div style='padding:6px 6px;text-align:center;white-space:normal;overflow-wrap:break-word;line-height:1.25;{mw}'>{col}</div></th>")
    html += "</tr></thead><tbody>"

    def is_total_row(row):
        for v in row.values:
            if isinstance(v, str) and ("합계" in v or "소계" in v or v.strip().startswith("총 ")):
                return True
        return False

    for _, row in d.iterrows():
        row_bold = is_total_row(row)
        html += "<tr>"
        for col in d.columns:
            val = row[col]
            if col in money_cols:
                try:
                    val_disp = f"{int(val):,}"
                except (TypeError, ValueError):
                    val_disp = fmt_plain(val)
                align = "right"
            else:
                val_disp = fmt_plain(val)
                align = "left" if col in left_cols else "center"
            has_html = "<" in val_disp
            mw = f"max-width:{col_max_width[col]};" if col in col_max_width else ""
            wrap_style = "white-space:normal;overflow:visible;text-overflow:clip;" if col in wrap_cols else "white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
            title_attr = "" if (has_html or col in wrap_cols) else f" title='{val_disp}'"
            bold_style = "font-weight:700;" if row_bold else ""
            html += (f"<td style='padding:0;border-bottom:1px solid #eee;'>"
                      f"<div style='padding:5px 10px;text-align:{align};{wrap_style}{mw}{bold_style}'{title_attr}>{val_disp}</div></td>")
        html += "</tr>"
    html += "</tbody></table></div>"
    st.markdown(html, unsafe_allow_html=True)


def is_overdue(current_due_date, status, today):
    if status in ("완납", "확인필요"):
        return False
    d = safe_date(current_due_date)
    return d is not None and d < today


def display_status(status, current_due_date, today):
    if status in ("완납", "확인필요"):
        return status
    overdue = is_overdue(current_due_date, status, today)
    if status == "일부입금":
        return "일부입금(지연)" if overdue else "일부입금"
    return "지연중" if overdue else "입금대기"


def calc_delay_days(original_due_date, ref_date):
    orig = safe_date(original_due_date)
    if orig is None or ref_date is None:
        return 0
    return max(0, (ref_date - orig).days)


def claim_severity(delay_count, delay_days):
    if delay_days >= 90:
        sev = 3
    elif delay_days >= 60:
        sev = 2
    elif delay_days >= 30:
        sev = 1
    else:
        sev = 0
    if delay_count >= 3:
        sev = max(sev, 1)
    return sev


SEVERITY_LABEL = {0: "정상", 1: "🟡 주의", 2: "🟠 경고", 3: "🔴 심각"}


# --------------------------------------------------------------------------
# 조회 캐시 — 안 보이는 탭도 스트림릿이 매번 다시 그리면서 조회를 반복하므로,
# 짧게(30초) 기억해뒀다가 재사용한다. 업로드 성공 시 clear_data_cache()로 즉시 비운다.
# --------------------------------------------------------------------------
@st.cache_data(ttl=30)
def load_table(_engine, table_name, where_clause=""):
    return pd.read_sql(f"SELECT * FROM {table_name} {where_clause};", _engine)


@st.cache_data(ttl=30)
def load_site_detail(_engine, site_id):
    return pd.read_sql(
        text("SELECT detail_type as 구분, detail_date as 일자, amount as 금액, note as 비고 "
             "FROM site_receivable_details WHERE site_receivable_id=:sid ORDER BY detail_date;"),
        _engine, params={"sid": site_id}
    )


def clear_data_cache():
    load_table.clear()
    load_site_detail.clear()


def run_daily_delay_check():
    today_str = date.today().isoformat()
    with engine.connect() as conn:
        open_claims = conn.execute(text("""
            SELECT id, current_due_date, original_due_date, last_flagged_due_date, status
            FROM claims WHERE status IN ('입금대기', '일부입금')
        """)).fetchall()
        for cid, cur_due, orig_due, last_flag, status in open_claims:
            due_d = safe_date(cur_due)
            if due_d is None or due_d >= date.today():
                continue
            if last_flag == cur_due:
                continue
            delay_days = calc_delay_days(orig_due, date.today())
            conn.execute(text("""
                INSERT INTO claim_delay_history (claim_id, event_type, old_due_date, new_due_date, delay_days, reason)
                VALUES (:cid, '자동지연', :due, :due, :ddays, '예정일 경과 자동 감지')
            """), {"cid": cid, "due": cur_due, "ddays": delay_days})
            conn.execute(text("UPDATE claims SET last_flagged_due_date=:due WHERE id=:cid"), {"due": cur_due, "cid": cid})
        conn.commit()


_today_str = date.today().isoformat()
if st.session_state.get("_delay_check_done_for") != _today_str:
    run_daily_delay_check()
    st.session_state["_delay_check_done_for"] = _today_str

st.sidebar.markdown("### 🔐 관리자 로그인")
pw_input = st.sidebar.text_input("관리자 비밀번호", type="password")
is_admin = pw_input == ADMIN_PASSWORD
if pw_input:
    if is_admin:
        st.sidebar.success("관리자 모드 ✅")
    else:
        st.sidebar.error("비밀번호 틀림")

tab_receivable, tab_progress, tab_calendar, tab_risk, tab_contract, tab_admin = st.tabs([
    "📋 현장별 미수현황", "📊 기성청구현황", "📅 입금 캘린더", "🚨 리스크 현장", "📈 계약현황", "🔐 관리자"
])

# ==========================================================================
# TAB: 현장별 미수현황  (미수내역 엑셀 미러링)
# ==========================================================================
with tab_receivable:
    st.subheader("📋 현장별 미수현황")
    st.caption("미수 잔액이 있는 현장 리스트입니다. (완납되면 목록에서 빠집니다)")
    sr_df = load_table(engine, "site_receivables", "ORDER BY contract_date")

    if sr_df.empty:
        st.info("데이터가 없습니다. '🔐 관리자' 탭에서 '현장별 미수관리' 엑셀을 업로드해주세요.")
    else:
        active_df = sr_df[sr_df["is_active"] == 1].copy()

        render_metric_cards([
            ("🏢", "활성 현장 수", f"{len(active_df)}개"),
            ("📄", "총 계약금액(부가세 포함)", fmt_eok(active_df["contract_amount"].sum())),
            ("💰", "총 입금액(부가세 포함)", fmt_eok(active_df["total_paid"].sum())),
            ("⚠️", "총 미수잔액(부가세 포함)", fmt_eok(active_df["unpaid_balance"].sum())),
        ])

        st.divider()
        st.markdown("#### 🔍 필터")
        f0, f1, f2, f3, f4 = st.columns(5)
        div_options = ["전체"] + sorted(active_df["division"].replace("", "미분류").fillna("미분류").unique().tolist())
        div_filter = f0.selectbox("구분별(ENC/필로브/대리점)", div_options)
        company_options = ["전체"] + sorted(active_df["company_name"].dropna().unique().tolist())
        company_filter = f1.selectbox("업체별", company_options)
        manager_options = ["전체"] + sorted(active_df["manager"].dropna().unique().tolist())
        manager_filter = f2.selectbox("담당자별", manager_options)
        year_series = pd.to_datetime(active_df["contract_date"], errors="coerce").dt.year.dropna().astype(int)
        year_options = ["전체"] + sorted(year_series.unique().tolist(), reverse=True)
        year_filter = f3.selectbox("계약일(연도)별", year_options)
        billing_filter = f4.selectbox("기성청구 필요 여부", ["전체", "필요한 현장만", "필요없음"])

        filtered_df = active_df.copy()
        div_series_all = filtered_df["division"].replace("", "미분류").fillna("미분류")
        if div_filter != "전체":
            filtered_df = filtered_df[div_series_all == div_filter]
        if company_filter != "전체":
            filtered_df = filtered_df[filtered_df["company_name"] == company_filter]
        if manager_filter != "전체":
            filtered_df = filtered_df[filtered_df["manager"] == manager_filter]
        if year_filter != "전체":
            filtered_df = filtered_df[pd.to_datetime(filtered_df["contract_date"], errors="coerce").dt.year == year_filter]

        prog_pct_all = (filtered_df["progress_rate"] * 100).round(0)
        inv_pct_all = (filtered_df["invoice_progress_rate"] * 100).round(0)
        billing_needed_all = (prog_pct_all >= 60) & (inv_pct_all < prog_pct_all)
        if billing_filter == "필요한 현장만":
            filtered_df = filtered_df[billing_needed_all]
        elif billing_filter == "필요없음":
            filtered_df = filtered_df[~billing_needed_all]

        st.caption(f"{len(filtered_df)}개 현장 표시 중")

        render_metric_cards([
            ("📄", "계약금액 합계(부가세 포함)", fmt_eok(filtered_df["contract_amount"].sum())),
            ("💰", "입금액 합계(부가세 포함)", fmt_eok(filtered_df["total_paid"].sum())),
            ("⚠️", "미수잔액 합계(부가세 포함)", fmt_eok(filtered_df["unpaid_balance"].sum())),
        ])

        st.divider()
        disp = filtered_df.rename(columns={
            "site_name": "현장명", "company_name": "업체명", "contract_date": "계약일",
            "contract_amount": "총계약금액", "total_paid": "총입금액", "unpaid_balance": "미수잔액",
            "manager": "담당자",
        })
        disp["번호"] = filtered_df["no_number"].apply(lambda x: str(int(x)) if pd.notna(x) else "-")
        disp["구분"] = filtered_df["division"].replace("", "미분류").fillna("미분류")
        disp["공정율(%)"] = (filtered_df["progress_rate"] * 100).round(0).astype(int)
        disp["기성율(%)"] = (filtered_df["invoice_progress_rate"] * 100).round(0).astype(int)
        def billing_note(row):
            prog, inv = row["공정율(%)"], row["기성율(%)"]
            if prog < 60 or inv >= prog:
                return ""
            if prog >= 80:
                return "<span style='color:#c0392b;'>잔금 청구 필요</span>"
            return "<span style='color:#c0392b;'>중도금 청구 필요</span>"

        disp["비고"] = disp.apply(billing_note, axis=1)
        cols = ["번호", "구분", "현장명", "업체명", "계약일", "총계약금액", "총입금액", "미수잔액", "공정율(%)", "기성율(%)", "담당자", "비고"]
        DIVISION_ORDER = {"ENC": 0, "필로브": 1, "대리점": 2, "해외": 3}
        show_df = (
            disp.assign(
                _sortno=filtered_df["no_number"],
                _sortdiv=disp["구분"].map(lambda x: DIVISION_ORDER.get(x, 99)),
            )
            .sort_values(["_sortdiv", "_sortno"])[cols]
            .reset_index(drop=True)
        )

        # 구분(ENC/필로브/대리점)이 바뀔 때마다 그 구분의 소계 행을 표 안에 끼워넣는다
        rows_with_subtotal = []
        prev_div = None
        group_rows = []

        def flush_subtotal(div_name, group):
            if not group:
                return
            gdf = pd.DataFrame(group)
            subtotal = {c: "" for c in cols}
            subtotal["현장명"] = f"▶ {div_name} 소계"
            subtotal["총계약금액"] = gdf["총계약금액"].sum()
            subtotal["총입금액"] = gdf["총입금액"].sum()
            subtotal["미수잔액"] = gdf["미수잔액"].sum()
            rows_with_subtotal.append(subtotal)

        for _, r in show_df.iterrows():
            cur_div = r["구분"]
            if prev_div is not None and cur_div != prev_div:
                flush_subtotal(prev_div, group_rows)
                group_rows = []
            group_rows.append(r[cols].to_dict())
            rows_with_subtotal.append(r[cols].to_dict())
            prev_div = cur_div
        flush_subtotal(prev_div, group_rows)

        grand_total = {c: "" for c in cols}
        grand_total["번호"] = f"총 {len(show_df)}개 현장"
        grand_total["총계약금액"] = show_df["총계약금액"].sum()
        grand_total["총입금액"] = show_df["총입금액"].sum()
        grand_total["미수잔액"] = show_df["미수잔액"].sum()
        rows_with_subtotal.append(grand_total)

        show_df_final = pd.DataFrame(rows_with_subtotal)[cols]
        render_html_table(show_df_final, money_cols=["총계약금액", "총입금액", "미수잔액"],
                           col_max_width={"비고": "120px"})

        st.markdown("## 🔍 현장 상세 내역")
        sel_site = st.selectbox("현장 선택", show_df["현장명"].tolist())
        st.write("")

        if sel_site:
            row = active_df[active_df["site_name"] == sel_site].iloc[0]

            summary_df = pd.DataFrame([{
                "현장명": row["site_name"], "업체명": row["company_name"],
                "계약일": row["contract_date"] or "-", "준공일": row["completion_date"] or "-",
                "착공일": row["start_date"] or "-",
                "총계약금(변경포함)": row["contract_amount"],
                "변경계약": row["change_amount"], "미수잔액": row["unpaid_balance"],
                "기성율(%)": round(row["invoice_progress_rate"] * 100),
                "담당자": row["manager"],
            }])
            render_html_table(summary_df, money_cols=["총계약금(변경포함)", "변경계약", "미수잔액"])

            render_progress_milestones(round(row["progress_rate"] * 100))

            detail_df = load_site_detail(engine, int(row["id"]))
            if detail_df.empty:
                st.caption("세부내역이 없습니다.")
            else:
                inv_df = detail_df[detail_df["구분"] == "계산서"].drop(columns=["구분", "비고"]).reset_index(drop=True)
                pay_df = detail_df[detail_df["구분"] == "입금"].drop(columns=["구분", "비고"]).reset_index(drop=True)
                change_df = detail_df[detail_df["구분"] == "변경계약"].drop(columns=["구분"]).reset_index(drop=True)

                def with_total_row(df):
                    total_row = pd.DataFrame([{"일자": "총합계", "금액": df["금액"].sum()}])
                    return pd.concat([df, total_row], ignore_index=True)

                st.divider()
                col_left, col_right = st.columns(2)
                with col_left:
                    st.markdown("**📄 계산서 발행 내역**")
                    if inv_df.empty:
                        st.caption("없음")
                    else:
                        render_html_table(with_total_row(inv_df), money_cols=["금액"], left_cols=[])
                with col_right:
                    st.markdown("**💰 입금 내역**")
                    if pay_df.empty:
                        st.caption("없음")
                    else:
                        render_html_table(with_total_row(pay_df), money_cols=["금액"], left_cols=[])

                if not change_df.empty:
                    st.markdown("**📝 변경계약 내역**")
                    render_html_table(change_df, money_cols=["금액"], left_cols=[])

# ==========================================================================
# TAB: 기성청구현황  (일일수금관리=이력 엑셀 기준)
# ==========================================================================
with tab_progress:
    st.subheader("📊 기성청구현황")
    st.caption("일일수금관리 기준 기성 청구 현황입니다. 현장별·담당자별로 입금예정일, 지연 여부 등을 확인할 수 있습니다.")
    view_mode = st.radio("보기 기준", ["현장별", "담당자별"], horizontal=True)
    st.divider()

    claims_df = load_table(engine, "claims")
    payments_df = load_table(engine, "payments")
    history_df = load_table(engine, "claim_delay_history")

    if claims_df.empty:
        st.info("데이터가 없습니다. '🔐 관리자' 탭에서 '일일수금관리' 엑셀을 업로드해주세요.")
    else:
        today = date.today()
        empty_df = pd.DataFrame()
        payments_by_claim = {cid: g for cid, g in payments_df.groupby("claim_id")} if not payments_df.empty else {}
        history_by_claim = {cid: g for cid, g in history_df.groupby("claim_id")} if not history_df.empty else {}

        claim_rows = []
        for _, c in claims_df.iterrows():
            cid = c["id"]
            pay_rows = payments_by_claim.get(cid, empty_df)
            paid = pay_rows["payment_amount"].sum() if not pay_rows.empty else 0
            unpaid = (c["claim_amount"] or 0) - paid
            hist_g = history_by_claim.get(cid, empty_df)
            hist = hist_g[hist_g["event_type"] == "자동지연"] if not hist_g.empty else empty_df
            delay_count = len(hist)
            if c["status"] == "완납":
                ref_date = (safe_date(pay_rows.iloc[-1]["payment_date"]) if not pay_rows.empty else today) or today
            else:
                ref_date = today
            delay_days = calc_delay_days(c["original_due_date"], ref_date) if c["status"] != "확인필요" else 0
            claim_rows.append({
                "id": cid, "현장명": c["site_name"], "업체명": c["company_name"] if pd.notna(c["company_name"]) else "-",
                "담당자": c["manager"], "채권종류": c["claim_type"],
                "최초예정일": c["original_due_date"], "입금예정일": c["current_due_date"],
                "청구금액": c["claim_amount"], "입금액": paid, "미수잔액": unpaid,
                "지연횟수": delay_count, "총지연일수": delay_days,
                "상태": display_status(c["status"], c["current_due_date"], today),
                "완납여부": c["status"] == "완납", "지연이력있음": delay_count >= 1,
            })
        claim_df_all = pd.DataFrame(claim_rows)

        if view_mode == "현장별":
            disp = claim_df_all[claim_df_all["미수잔액"] > 0].copy()
            f1, f2, f3, f4 = st.columns(4)
            site_filter = f1.selectbox("현장 필터", ["전체"] + sorted(disp["현장명"].unique().tolist()))
            type_filter = f2.selectbox("채권종류 필터", ["전체"] + CLAIM_TYPES)
            manager_filter = f3.selectbox("담당자 필터", ["전체"] + sorted(disp["담당자"].dropna().unique().tolist()))
            company_filter = f4.selectbox("업체명 필터", ["전체"] + sorted(disp["업체명"].dropna().unique().tolist()))
            if site_filter != "전체":
                disp = disp[disp["현장명"] == site_filter]
            if type_filter != "전체":
                disp = disp[disp["채권종류"] == type_filter]
            if manager_filter != "전체":
                disp = disp[disp["담당자"] == manager_filter]
            if company_filter != "전체":
                disp = disp[disp["업체명"] == company_filter]

            disp = disp.assign(_sort=pd.to_datetime(disp["최초예정일"], errors="coerce")).sort_values("_sort", na_position="last")

            def colorize_status(s):
                if s in ("지연중", "일부입금(지연)"):
                    return f"<span style='background:#FCEBEB;color:#A32D2D;border-radius:999px;padding:2px 10px;font-size:12px;'>{s}</span>"
                if s == "확인필요":
                    return f"<span style='background:#FAEEDA;color:#854F0B;border-radius:999px;padding:2px 10px;font-size:12px;'>{s}</span>"
                return f"<span style='background:#E6F1FB;color:#185FA5;border-radius:999px;padding:2px 10px;font-size:12px;'>{s}</span>"

            cols_bysite = ["현장명", "업체명", "담당자", "채권종류", "최초예정일", "입금예정일", "청구금액", "입금액", "미수잔액", "지연횟수", "총지연일수", "상태"]
            show = disp[cols_bysite]

            total_row = {c: "" for c in cols_bysite}
            total_row["현장명"] = f"총 {len(show)}건"
            total_row["청구금액"] = show["청구금액"].sum()
            total_row["입금액"] = show["입금액"].sum()
            total_row["미수잔액"] = show["미수잔액"].sum()
            show_with_total = pd.concat([show, pd.DataFrame([total_row])], ignore_index=True)

            show_display = show_with_total.copy()
            show_display["상태"] = show_display["상태"].apply(colorize_status)
            render_html_table(show_display, money_cols=["청구금액", "입금액", "미수잔액"])
            st.download_button("📥 CSV 다운로드", show.to_csv(index=False).encode("utf-8-sig"),
                                file_name=f"기성청구현황_현장별_{date.today()}.csv", mime="text/csv")

        else:  # 담당자별
            managers = sorted(m for m in claim_df_all["담당자"].dropna().unique().tolist()
                               if str(m).strip() and str(m).strip().lower() != "nan")
            rows = []
            for m in managers:
                mc = claim_df_all[claim_df_all["담당자"] == m]
                total_cnt = len(mc)
                completed_cnt = int(mc["완납여부"].sum())
                unpaid_cnt = total_cnt - completed_cnt
                ever_delayed_cnt = int(mc["지연이력있음"].sum())
                total_claim = mc["청구금액"].sum()
                total_paid = mc["입금액"].sum()
                unpaid_amt = mc["미수잔액"].sum()
                rows.append({
                    "담당자": m, "총청구금액": total_claim, "총입금액": total_paid, "미수잔액": unpaid_amt,
                    "총청구건": total_cnt, "완납건": completed_cnt, "미수건": unpaid_cnt, "지연건": ever_delayed_cnt,
                    "수금률(%)": round(total_paid / total_claim * 100, 1) if total_claim else 0,
                    "지연률(%)": round(ever_delayed_cnt / total_cnt * 100, 1) if total_cnt else 0,
                })
            manager_df = pd.DataFrame(rows)
            render_html_table(manager_df, money_cols=["총청구금액", "총입금액", "미수잔액"])
            st.download_button("📥 CSV 다운로드", manager_df.to_csv(index=False).encode("utf-8-sig"),
                                file_name=f"기성청구현황_담당자별_{date.today()}.csv", mime="text/csv")

            st.divider()
            st.markdown("#### 🔍 담당자 상세 (채권종류별 3단 현황)")
            pick_m = st.selectbox("담당자 선택", managers)
            mc = claim_df_all[claim_df_all["담당자"] == pick_m]

            st.markdown("**1. 청구현황** (채권종류별 전체 건수·금액)")
            t1 = mc.groupby("채권종류").agg(건수=("id", "count"), 금액=("청구금액", "sum")).reset_index()
            t1 = pd.concat([t1, pd.DataFrame([{"채권종류": "합계", "건수": t1["건수"].sum(), "금액": t1["금액"].sum()}])], ignore_index=True)
            render_html_table(t1, money_cols=["금액"], left_cols=[])

            st.markdown("**2. 미수현황** (채권종류별 미수 건수·금액)")
            mc_unpaid = mc[mc["미수잔액"] > 0]
            if mc_unpaid.empty:
                st.caption("미수 없음")
            else:
                t2 = mc_unpaid.groupby("채권종류").agg(건수=("id", "count"), 미수금액=("미수잔액", "sum")).reset_index()
                t2 = pd.concat([t2, pd.DataFrame([{"채권종류": "합계", "건수": t2["건수"].sum(), "미수금액": t2["미수금액"].sum()}])], ignore_index=True)
                render_html_table(t2, money_cols=["미수금액"], left_cols=[])

            st.markdown("**3. 지연현황** (채권종류별 지연 건수·총지연일수)")
            mc_delayed = mc[mc["지연이력있음"]]
            if mc_delayed.empty:
                st.caption("지연 이력 없음")
            else:
                t3 = mc_delayed.groupby("채권종류").agg(건수=("id", "count"), 총지연일수=("총지연일수", "sum")).reset_index()
                t3 = pd.concat([t3, pd.DataFrame([{"채권종류": "합계", "건수": t3["건수"].sum(), "총지연일수": t3["총지연일수"].sum()}])], ignore_index=True)
                render_html_table(t3, money_cols=[], left_cols=[])

# ==========================================================================
# TAB: 입금 캘린더 (이력 데이터만)
# ==========================================================================
with tab_calendar:
    st.subheader("📅 입금 캘린더")
    st.caption("기성 청구된 입금예정 건을 달력으로 확인합니다. 완납(초록)·지연(빨강)·입금대기(회색)로 구분됩니다.")
    claims_df = load_table(
        engine, "claims",
        "WHERE (current_due_date IS NOT NULL AND current_due_date != '') "
        "OR (original_due_date IS NOT NULL AND original_due_date != '')"
    )
    history_df = load_table(engine, "claim_delay_history")
    checkpoints_df = load_table(engine, "claim_checkpoints")
    payments_df = load_table(engine, "payments")

    if claims_df.empty:
        st.info("데이터가 없습니다.")
    else:
        today = date.today()
        if "cal_year" not in st.session_state:
            st.session_state.cal_year = today.year
            st.session_state.cal_month = today.month

        nav1, nav2, nav3 = st.columns([1, 3, 1])
        if nav1.button("◀ 이전달", use_container_width=True):
            st.session_state.cal_month -= 1
            if st.session_state.cal_month == 0:
                st.session_state.cal_month = 12
                st.session_state.cal_year -= 1
            st.rerun()
        nav2.markdown(f"<h2 style='text-align:center;margin:0;'>{st.session_state.cal_year}년 {st.session_state.cal_month}월</h2>", unsafe_allow_html=True)
        if nav3.button("다음달 ▶", use_container_width=True):
            st.session_state.cal_month += 1
            if st.session_state.cal_month == 13:
                st.session_state.cal_month = 1
                st.session_state.cal_year += 1
            st.rerun()

        st.caption("🔴 지연　🟢 입금완료　⚪ 입금대기")
        yr, mo = st.session_state.cal_year, st.session_state.cal_month

        # 청구건마다 매번 전체 표를 다시 훑으면(예: checkpoints_df[checkpoints_df["claim_id"]==cid]) 느려지므로,
        # claim_id 기준으로 미리 한 번만 묶어두고 그 후엔 딕셔너리로 바로 찾아 쓴다
        checkpoints_by_claim = {cid: g for cid, g in checkpoints_df.groupby("claim_id")} if not checkpoints_df.empty else {}
        history_by_claim = {cid: g for cid, g in history_df.groupby("claim_id")} if not history_df.empty else {}
        payments_by_claim = {cid: g for cid, g in payments_df.groupby("claim_id")} if not payments_df.empty else {}
        empty_df = pd.DataFrame()

        day_entries = {}
        for _, c in claims_df.iterrows():
            cid = c["id"]
            site_short = c["site_name"][:14] + ("…" if len(c["site_name"]) > 14 else "")
            amt_disp = f"{int(c['claim_amount']) // 1000:,}"

            cur_d = safe_date(c["current_due_date"])

            # 1) 현재(최종) 예정일 — 항상 실시간 상태로 표시
            if cur_d and cur_d.year == yr and cur_d.month == mo:
                st_disp = display_status(c["status"], c["current_due_date"], today)
                if c["status"] == "완납":
                    color = "#27ae60"
                elif "지연" in st_disp:
                    color = "#e74c3c"
                else:
                    color = "#95a5a6"
                day_entries.setdefault(cur_d.day, []).append(
                    {"site": site_short, "claim_type": c["claim_type"], "amt": amt_disp, "color": color}
                )

            # 2) 과거에 밀렸던 예정일들 전부 — 그 시점(체크포인트)에 실제로 미수였던 날짜는
            #    지금 완납됐어도 그날짜엔 그대로 남긴다 (그 순간엔 진짜 미수였으니까)
            cp_rows = checkpoints_by_claim.get(cid, empty_df)
            for _, cp in cp_rows.iterrows():
                cp_d = safe_date(cp["checkpoint_date"])
                if not cp_d or cp_d.year != yr or cp_d.month != mo or cp_d >= today or cp_d == cur_d:
                    continue
                cp_unpaid = cp["unpaid_balance"]
                if cp_unpaid is not None and cp_unpaid > 0:
                    day_entries.setdefault(cp_d.day, []).append(
                        {"site": site_short, "claim_type": c["claim_type"] + "(경과된 예정일)", "amt": amt_disp, "color": "#e74c3c"}
                    )

        cal = pycal.Calendar(firstweekday=6)
        weeks = cal.monthdayscalendar(yr, mo)
        weekday_labels = ["일", "월", "화", "수", "목", "금", "토"]

        MAX_SHOWN = 4
        BODY_HEIGHT = 92  # 버튼(날짜) 아래 내용 박스의 고정 높이 — 내용 없어도 항상 이 높이로 고정

        st.markdown("""
        <style>
        div[data-testid="stHorizontalBlock"] div.stButton > button {
            border-radius: 6px 6px 0 0 !important;
            border: none !important;
            font-weight: 800 !important;
            padding: 2px 0 !important;
        }
        div[data-testid="stHorizontalBlock"] div.stButton > button[kind="secondary"] {
            background-color: #eeeeee !important;
        }
        </style>
        """, unsafe_allow_html=True)

        header_cols = st.columns(7, gap="small")
        for i, lab in enumerate(weekday_labels):
            wcolor = "#ff8080" if i == 0 else ("#8ab4ff" if i == 6 else "#fff")
            header_cols[i].markdown(
                f"<div style='text-align:center;font-weight:800;padding:6px 0;background:#333;color:{wcolor};border-radius:6px;'>{lab}</div>",
                unsafe_allow_html=True
            )

        for week in weeks:
            row_cols = st.columns(7, gap="small")
            for i, daynum in enumerate(week):
                with row_cols[i]:
                    if daynum == 0:
                        st.markdown(
                            f"<div style='height:{BODY_HEIGHT + 38}px;background:#fafafa;border-radius:6px;'></div>",
                            unsafe_allow_html=True
                        )
                        continue

                    is_today = (daynum == today.day and yr == today.year and mo == today.month)

                    if st.button(str(daynum), key=f"cal_{yr}_{mo}_{daynum}", use_container_width=True,
                                 type="primary" if is_today else "secondary"):
                        st.session_state["cal_selected_date"] = date(yr, mo, daynum).isoformat()

                    entries = day_entries.get(daynum, [])
                    body = ""
                    for e in entries[:MAX_SHOWN]:
                        body += (
                            f"<div style='font-size:11px;line-height:1.35;color:#000;white-space:nowrap;overflow:hidden;"
                            f"text-overflow:ellipsis;' title='{e['site']} {e['claim_type']} {e['amt']}'>"
                            f"<span style='color:{e['color']};font-size:13px;'>●</span> {e['site']}</div>"
                        )
                    if len(entries) > MAX_SHOWN:
                        body += f"<div style='font-size:10px;color:#999;'>+{len(entries) - MAX_SHOWN}건 더</div>"

                    # 테두리 없이 옅은 배경만 깔아서, 버튼과의 이음새가 안 보이게 함
                    st.markdown(
                        f"<div style='height:{BODY_HEIGHT}px;overflow:hidden;background:#f7f7f7;"
                        f"border-radius:0 0 6px 6px;padding:4px 5px;margin-top:-6px;'>{body}</div>",
                        unsafe_allow_html=True
                    )
            st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        st.divider()
        dates_with_entries = sorted(f"{yr}-{mo:02d}-{d:02d}" for d in day_entries.keys())
        if dates_with_entries:
            clicked = st.session_state.get("cal_selected_date")
            default_idx = dates_with_entries.index(clicked) if clicked in dates_with_entries else 0
            sel_date = st.selectbox("📌 날짜별 상세 목록 보기 (달력에서 날짜를 눌러도 여기로 이동합니다)", dates_with_entries, index=default_idx)
            sel_d = safe_date(sel_date)

            day_rows = []
            for _, c in claims_df.iterrows():
                cid = c["id"]
                cur_d = safe_date(c["current_due_date"])
                hist_g = history_by_claim.get(cid, empty_df)
                hist_all = hist_g[hist_g["event_type"] == "자동지연"] if not hist_g.empty else empty_df

                cp_g = checkpoints_by_claim.get(cid, empty_df)
                cp_today = cp_g[cp_g["checkpoint_date"] == sel_date] if not cp_g.empty else empty_df

                is_current_match = (cur_d == sel_d)
                is_checkpoint_match = False
                if not is_current_match and not cp_today.empty and sel_d < today:
                    cp_unpaid = cp_today.iloc[-1]["unpaid_balance"]
                    if cp_unpaid is not None and cp_unpaid > 0:
                        is_checkpoint_match = True

                if not (is_current_match or is_checkpoint_match):
                    continue

                delay_count = len(hist_all)
                delay_days = calc_delay_days(c["original_due_date"], today) if c["status"] != "확인필요" else 0

                if is_checkpoint_match:
                    status_label, sort_rank = "지연중", 1
                elif c["status"] == "완납":
                    status_label, sort_rank = "완납", 0
                elif c["status"] == "확인필요":
                    status_label, sort_rank = "확인필요", 3
                elif c["status"] == "일부입금":
                    status_label, sort_rank = ("일부입금(지연)", 1) if delay_days > 0 else ("일부입금", 2)
                else:
                    status_label, sort_rank = ("지연중", 1) if delay_days > 0 else ("입금대기", 2)

                if status_label == "완납":
                    status_html = f"<span style='background:#EAF3DE;color:#27500A;border-radius:999px;padding:2px 10px;font-size:12px;'>{status_label}</span>"
                elif "지연" in status_label:
                    status_html = f"<span style='background:#FCEBEB;color:#A32D2D;border-radius:999px;padding:2px 10px;font-size:12px;'>{status_label}</span>"
                else:
                    status_html = f"<span style='background:#E6F1FB;color:#185FA5;border-radius:999px;padding:2px 10px;font-size:12px;'>{status_label}</span>"

                # 그 날짜(sel_d)에 실제로 적혀있던 비고를 그대로 찾아서 붙인다 (최신 비고로 통일하지 않음)
                remark_v = ""
                if not cp_today.empty:
                    remark_v = cp_today.iloc[-1]["remark"] or ""
                if not remark_v and is_current_match:
                    remark_v = c["last_remark"] if pd.notna(c["last_remark"]) else ""
                if str(remark_v).strip().lower() in ("nan", "none"):
                    remark_v = ""

                pay_rows = payments_by_claim.get(cid, empty_df)
                paid_date_v = pay_rows["payment_date"].max() if not pay_rows.empty else "-"

                day_rows.append({
                    "현장명": c["site_name"], "업체명": c["company_name"] if pd.notna(c["company_name"]) else "-",
                    "채권종류": c["claim_type"], "담당자": c["manager"] if pd.notna(c["manager"]) else "-",
                    "청구금액": c["claim_amount"],
                    "상태": status_html, "최초예정일": c["original_due_date"] or "-", "입금일": paid_date_v,
                    "지연횟수": delay_count, "총지연일수": delay_days,
                    "비고": remark_v,
                    "_sort": sort_rank, "_status_plain": status_label,
                })
            day_df_raw = pd.DataFrame(day_rows)
            pending_total = day_df_raw[day_df_raw["_status_plain"] != "완납"]["청구금액"].sum() if not day_df_raw.empty else 0
            day_df = day_df_raw.sort_values("_sort").drop(columns=["_sort", "_status_plain"]) if day_rows else day_df_raw
            render_html_table(day_df, money_cols=["청구금액"], wrap_cols=["현장명", "비고"], left_cols=["현장명", "비고"])
            st.metric("💰 이 날짜 미수금액 합계 (완납 제외)", f"{pending_total:,.0f} 원")
        else:
            st.caption("이 달에는 예정된 청구가 없습니다.")

# ==========================================================================
# TAB: 리스크 현장 (이력 데이터만)
# ==========================================================================
with tab_risk:
    st.subheader("🚨 리스크 현장")
    st.caption("완납되지 않은 청구 중 지연 3회 이상이거나 지연일수 30일 이상인 건이 있는 현장. (완납 청구는 제외)")
    claims_df = load_table(engine, "claims", "WHERE status != '완납'")
    history_df = load_table(engine, "claim_delay_history")

    if claims_df.empty:
        st.info("데이터가 없습니다.")
    else:
        today = date.today()
        empty_df = pd.DataFrame()
        history_by_claim = {cid: g for cid, g in history_df.groupby("claim_id")} if not history_df.empty else {}
        risk_rows = []
        for _, c in claims_df.iterrows():
            cid = c["id"]
            hist_g = history_by_claim.get(cid, empty_df)
            hist = hist_g[hist_g["event_type"] == "자동지연"] if not hist_g.empty else empty_df
            delay_count = len(hist)
            delay_days = calc_delay_days(c["original_due_date"], today) if c["status"] != "확인필요" else 0
            sev = claim_severity(delay_count, delay_days)
            if sev == 0:
                continue
            reasons = []
            if delay_count >= 3:
                reasons.append(f"지연 {delay_count}회")
            if delay_days >= 30:
                reasons.append(f"{delay_days}일 지연")
            risk_rows.append({
                "현장명": c["site_name"], "업체명": c["company_name"] if pd.notna(c["company_name"]) else "-",
                "담당자": c["manager"], "채권종류": c["claim_type"],
                "청구금액": c["claim_amount"], "최초예정일": c["original_due_date"], "입금예정일": c["current_due_date"],
                "지연횟수": delay_count, "지연일수": delay_days, "등급": SEVERITY_LABEL[sev], "_sev": sev,
                "사유": ", ".join(reasons),
            })

        if not risk_rows:
            st.success("현재 리스크 현장이 없습니다. 👍")
        else:
            risk_df = pd.DataFrame(risk_rows)

            render_metric_cards([
                ("🏢", "리스크 현장 수", f"{risk_df['현장명'].nunique()}개"),
                ("🔴", "심각 등급 현장", f"{(risk_df.groupby('현장명')['_sev'].max() == 3).sum()}개"),
                ("🧾", "리스크 청구 건수", f"{len(risk_df)}건"),
            ])

            st.divider()
            st.markdown("#### 리스크 청구 상세")
            display_risk = risk_df.sort_values(["_sev", "지연일수"], ascending=[False, False]).copy()

            SEV_PLAIN = {1: "주의", 2: "경고", 3: "심각"}
            SEV_BADGE = {
                1: "<span style='background:#F7C1C1;color:#791F1F;border-radius:999px;padding:2px 10px;font-size:12px;'>주의</span>",
                2: "<span style='background:#E24B4A;color:#fff;border-radius:999px;padding:2px 10px;font-size:12px;'>경고</span>",
                3: "<span style='background:#791F1F;color:#fff;border-radius:999px;padding:2px 10px;font-size:12px;'>심각</span>",
            }
            csv_export = display_risk.copy()
            csv_export["등급"] = csv_export["_sev"].map(SEV_PLAIN)
            csv_export = csv_export.drop(columns=["_sev"])

            display_risk["등급"] = display_risk["_sev"].map(SEV_BADGE)
            display_risk = display_risk.drop(columns=["_sev"])

            render_html_table(display_risk, money_cols=["청구금액"])
            st.download_button("📥 CSV 다운로드", csv_export.to_csv(index=False).encode("utf-8-sig"),
                                file_name=f"리스크현장_{date.today()}.csv", mime="text/csv")

# ==========================================================================
# TAB: 계약현황 (미수내역 전체 - 활성/완불 구분 없이 계약년월 있는 모든 행)
# ==========================================================================
with tab_contract:
    st.subheader("📈 계약현황 (연도별)")
    cs_df = load_table(engine, "contract_status_raw")

    if cs_df.empty:
        st.info("데이터가 없습니다. '🔐 관리자' 탭에서 '현장별 미수관리' 엑셀을 업로드해주세요.")
    else:
        LABEL_MAP = {
            "총 계약금": "총계약금", "서울": "서울계약", "대구": "대구계약", "대리점": "대리점계약", "해외": "해외계약",
            "총 현장수": "현장수", "서울현장수": "서울현장수", "대구현장수": "대구현장수",
            "대리점현장수": "대리점현장수", "해외현장수": "해외현장수",
        }
        cs_df = cs_df[cs_df["label"].isin(LABEL_MAP.keys())].copy()
        cs_df["label"] = cs_df["label"].map(LABEL_MAP)

        SITE_COUNT_COLS = ["현장수", "서울현장수", "대구현장수", "대리점현장수", "해외현장수"]

        yearly = cs_df.pivot_table(index="year", columns="label", values="total", aggfunc="first").reset_index()
        yearly = yearly.rename(columns={"year": "연도"})
        yearly["연도"] = yearly["연도"].astype(int)
        col_order = ["연도", "총계약금", "현장수", "서울계약", "서울현장수", "대구계약", "대구현장수",
                     "대리점계약", "대리점현장수", "해외계약", "해외현장수"]
        col_order = [c for c in col_order if c in yearly.columns]
        yearly = yearly[col_order].sort_values("연도")
        for c in SITE_COUNT_COLS:
            if c in yearly.columns:
                yearly[c] = yearly[c].fillna(0).astype(int)

        money_cols = [c for c in yearly.columns if "계약" in c]
        render_html_table(yearly, money_cols=money_cols, left_cols=[], fixed_layout=True,
                           narrow_cols=SITE_COUNT_COLS)

        st.divider()
        sel_year = st.selectbox("월별로 보기", sorted(cs_df["year"].unique().tolist(), reverse=True))
        month_cols = [f"m{i}" for i in range(1, 13)]
        y_df = cs_df[cs_df["year"] == sel_year].set_index("label")[month_cols].T
        y_df.index = range(1, 13)
        y_df = y_df.reset_index().rename(columns={"index": "월"})
        col_order_m = ["월", "총계약금", "현장수", "서울계약", "서울현장수", "대구계약", "대구현장수",
                       "대리점계약", "대리점현장수", "해외계약", "해외현장수"]
        col_order_m = [c for c in col_order_m if c in y_df.columns]
        y_df = y_df[col_order_m]
        for c in SITE_COUNT_COLS:
            if c in y_df.columns:
                y_df[c] = y_df[c].fillna(0).astype(int)
        money_cols_m = [c for c in y_df.columns if "계약" in c]
        render_html_table(y_df, money_cols=money_cols_m, left_cols=[], fixed_layout=True,
                           narrow_cols=SITE_COUNT_COLS)

        st.download_button("📥 연도별 CSV 다운로드", yearly.to_csv(index=False).encode("utf-8-sig"),
                            file_name=f"계약현황_연도별_{date.today()}.csv", mime="text/csv")

        st.divider()
        st.markdown("## 📊 기간별 전년대비 비교")
        st.caption("시작월, 종료월을 고르면 매년 기간별 누적 계약금액과 현장수를 비교하고, 전년대비 증감을 보여줍니다.")

        pc1, pc2, pc3 = st.columns(3)
        m_start = pc1.selectbox("시작월", list(range(1, 13)), index=0)
        m_end = pc2.selectbox("종료월", list(range(1, 13)), index=date.today().month - 1)
        region_options = ["전체", "서울", "대구", "대리점", "해외"]
        region_filter = pc3.selectbox("지역 필터", region_options)
        REGION_LABEL_MAP = {
            "전체": ("총계약금", "현장수"),
            "서울": ("서울계약", "서울현장수"),
            "대구": ("대구계약", "대구현장수"),
            "대리점": ("대리점계약", "대리점현장수"),
            "해외": ("해외계약", "해외현장수"),
        }

        if m_start > m_end:
            st.warning("시작월이 종료월보다 늦습니다. 순서를 바꿔주세요.")
        else:
            amt_label, cnt_label = REGION_LABEL_MAP[region_filter]
            months_in_range = [f"m{i}" for i in range(m_start, m_end + 1)]
            years_all = sorted(cs_df["year"].unique().tolist())
            comp_rows = []
            for yr_c in years_all:
                yr_data = cs_df[cs_df["year"] == yr_c]
                total_amt = 0
                site_cnt = 0
                for _, r in yr_data.iterrows():
                    vals = [r[m] for m in months_in_range if pd.notna(r[m])]
                    if r["label"] == amt_label:
                        total_amt = sum(vals)
                    elif r["label"] == cnt_label:
                        site_cnt = int(sum(vals))
                comp_rows.append({"연도": yr_c, "누적계약금액": total_amt, "현장수": site_cnt})

            comp_df = pd.DataFrame(comp_rows).sort_values("연도")
            amt_col = "누적계약금액"
            comp_df["전년대비 증감액"] = comp_df[amt_col].diff()
            raw_pct = comp_df[amt_col].pct_change() * 100
            raw_pct = raw_pct.replace([float("inf"), float("-inf")], float("nan"))  # 전년도 값이 0이면 증감률 계산 불가 -> 공란
            comp_df["전년대비 증감률(%)"] = raw_pct.round(1)
            comp_df["전년대비 증감액"] = comp_df["전년대비 증감액"].fillna(0).astype(int)

            render_html_table(comp_df, money_cols=[amt_col, "전년대비 증감액"], left_cols=[])

            chart_data = comp_df[["연도", amt_col]].copy()
            chart_data["연도"] = chart_data["연도"].astype(str)
            chart_data["억원표시"] = (chart_data[amt_col] / 1e8).round(0).astype(int).astype(str) + "억"
            axis_expr = "format(datum.value/100000000, ',.0f') + '억'"
            UNIT = 1_000_000_000  # 10억
            max_val = float(chart_data[amt_col].max()) if not chart_data.empty else 0
            n_ticks = int(max_val // UNIT) + 2
            tick_values = [i * UNIT for i in range(n_ticks)]

            bar_chart = (
                alt.Chart(chart_data)
                .mark_bar(size=28, color="#4C78A8", tooltip=False)
                .encode(
                    x=alt.X("연도:N", sort=None, title="연도", axis=alt.Axis(labelAngle=0)),
                    y=alt.Y(f"{amt_col}:Q", title="금액(억원)",
                            axis=alt.Axis(labelExpr=axis_expr, values=tick_values, labelOverlap=False)),
                )
            )
            right_axis_layer = (
                alt.Chart(chart_data)
                .mark_bar(size=28, opacity=0, tooltip=False)
                .encode(
                    x=alt.X("연도:N", sort=None),
                    y=alt.Y(f"{amt_col}:Q", title=None,
                            axis=alt.Axis(orient="right", labelExpr=axis_expr, values=tick_values, labelOverlap=False)),
                )
            )
            combined_chart = (
                alt.layer(bar_chart, right_axis_layer)
                .resolve_scale(y="shared")
                .resolve_axis(y="independent")
                .properties(height=630)
            )
            st.altair_chart(combined_chart, use_container_width=True)



# ==========================================================================
# TAB: 관리자 — 엑셀 업로드 2개만
# ==========================================================================
with tab_admin:
    st.subheader("🔐 관리자")
    if not is_admin:
        st.warning("🔒 관리자 전용 메뉴입니다. 왼쪽 사이드바에서 비밀번호를 입력하세요.")
    else:
        st.markdown("업로드하면 **해당 데이터 전체가 지금 올리는 파일 내용으로 갈음**됩니다 (기존 것 삭제 후 새로 채움).")

        # ---------------- 일일수금관리 업로드 ----------------
        st.markdown("### 📂 일일수금관리 업로드 (기성청구현황·캘린더·리스크현장)")
        daily_file = st.file_uploader("일일수금관리 엑셀(.xlsx) 업로드", type=["xlsx", "csv"], key="daily_upload")

        if daily_file is not None and st.button("🚀 일일수금관리 데이터로 전체 갱신", use_container_width=True):
            try:
                xls = pd.ExcelFile(daily_file, engine="openpyxl")
                if "이력" not in xls.sheet_names:
                    st.error("'이력' 시트를 찾지 못했습니다.")
                    st.stop()
                raw_df = pd.read_excel(xls, sheet_name="이력", header=1)
                raw_df.columns = [str(c).strip() for c in raw_df.columns]

                col_map = {
                    "채권명": "bond_name", "현장명": "site_name", "업체명": "company_name",
                    "담당자": "manager", "기성구분": "claim_type", "미수금액": "claim_amount",
                    "최초예정일": "original_due_date", "입금예정일": "current_due_date",
                    "지연여부": "status_raw", "입금여부": "paid_flag", "입금액": "paid_amount",
                    "비고": "remark", "미수잔액": "sheet_unpaid_balance",
                }
                for c in list(raw_df.columns):
                    if "입금상태" in c or "입금일자" in c:
                        col_map[c] = "payment_date_raw"
                raw_df = raw_df.rename(columns={k: v for k, v in col_map.items() if k in raw_df.columns})
                raw_df = raw_df.dropna(subset=["bond_name"]) if "bond_name" in raw_df.columns else raw_df.dropna(subset=["site_name"])
                raw_df["_due_sort"] = pd.to_datetime(raw_df.get("current_due_date"), errors="coerce")

                with engine.connect() as conn:
                    for t in ["claim_delay_history", "claim_checkpoints", "payments", "claims"]:
                        conn.execute(text(f"DELETE FROM {t};"))
                    conn.commit()

                    has_bond = "bond_name" in raw_df.columns
                    groups = raw_df.groupby("bond_name") if has_bond else [(i, raw_df.loc[[i]]) for i in raw_df.index]
                    n_claims = 0
                    claims_batch = []
                    history_batch = []
                    checkpoint_batch = []
                    payment_batch = []

                    for _, grp_raw in groups:
                        # 채권명이 같아도(현장+업체+구분+금액 우연히 동일), 중간에 미수잔액이 0으로
                        # 완전히 끝났다가 이후에 다시 미수금이 생기면 별개의 새 채권으로 취급해서 쪼갠다.
                        # 이때 반드시 '진짜 시간 순서'(입금예정일, 같으면 입금완료 행을 나중으로)로 봐야 한다 —
                        # 엑셀에서 사용자가 행을 정렬/이동하면 물리적 위치는 시간 순서와 달라질 수 있기 때문.
                        grp_tmp = grp_raw.copy()
                        grp_tmp["_paid_flag_priority"] = (grp_tmp.get("paid_flag", "").astype(str).str.upper() == "Y").astype(int)
                        grp_phys = grp_tmp.sort_values(["_due_sort", "_paid_flag_priority"], na_position="last")
                        episodes = []
                        cur_ep = []
                        for _, r in grp_phys.iterrows():
                            cur_ep.append(r)
                            rb = r.get("sheet_unpaid_balance", None)
                            try:
                                rb_val = float(rb) if rb is not None and pd.notna(rb) else None
                            except (TypeError, ValueError):
                                rb_val = None
                            if rb_val is not None and rb_val <= 0:
                                episodes.append(pd.DataFrame(cur_ep))
                                cur_ep = []
                        if cur_ep:
                            episodes.append(pd.DataFrame(cur_ep))
                        if not episodes:
                            continue

                        for grp in episodes:
                            grp = grp.sort_values("_due_sort", na_position="last")
                            first, last = grp.iloc[0], grp.iloc[-1]
                            site_name = str(first.get("site_name", "")).strip()
                            company_raw = first.get("company_name", "")
                            company_name_v = "" if pd.isna(company_raw) else str(company_raw).strip()
                            manager_raw = first.get("manager", "")
                            manager = "" if pd.isna(manager_raw) else str(manager_raw).strip()
                            claim_type_raw = last.get("claim_type", "")
                            claim_type_v = "기성금" if pd.isna(claim_type_raw) else (str(claim_type_raw).strip() or "기성금")
                            claim_amount = parse_amount(last.get("claim_amount", 0))
                            orig_due = safe_date(first.get("original_due_date")) or safe_date(first.get("current_due_date"))
                            orig_due_str = orig_due.isoformat() if orig_due else None

                            valid_date_rows = grp[grp["_due_sort"].notna()]
                            last_valid_row = valid_date_rows.iloc[-1] if not valid_date_rows.empty else last
                            cur_due = safe_date(last_valid_row.get("current_due_date"))
                            cur_due_str = cur_due.isoformat() if cur_due else None
                            status_raw = str(last.get("status_raw", "") or "")

                            payment_events = []
                            for _, r in grp.iterrows():
                                pa = parse_amount(r.get("paid_amount", 0))
                                pdate = safe_date(r.get("payment_date_raw"))
                                if pdate and pa > 0:
                                    payment_events.append((pdate, pa))

                            # 이력 로그를 그냥 다 더하면, 나중에 취소/정정된 입금까지 합산돼서 완납으로 잘못 판정될 수 있다.
                            # 그래서 "지금 이 순간 미수잔액이 얼마인가"를 이력의 마지막 줄에서 직접 읽어와 그걸 진실로 삼는다.
                            sheet_unpaid_raw = last_valid_row.get("sheet_unpaid_balance", None)
                            if pd.notna(sheet_unpaid_raw):
                                sheet_unpaid = parse_amount(sheet_unpaid_raw)
                                total_paid = max(0, claim_amount - sheet_unpaid)
                            else:
                                total_paid = sum(a for _, a in payment_events)

                            if total_paid >= claim_amount and claim_amount > 0:
                                status = "완납"
                            elif total_paid > 0:
                                status = "일부입금"
                            elif cur_due_str is None:
                                status = "확인필요"
                            else:
                                status = "입금대기"

                            remark_raw = last.get("remark", "")
                            remark_v = "" if pd.isna(remark_raw) else str(remark_raw).strip()
                            if remark_v.lower() in ("none", "nan"):
                                remark_v = ""

                            claims_batch.append({
                                "sn": site_name, "cn": company_name_v, "mg": manager, "ctype": claim_type_v,
                                "cdate": orig_due_str, "odue": orig_due_str, "cdue": cur_due_str,
                                "amt": claim_amount, "status": status, "remark": remark_v,
                            })
                            claim_idx = len(claims_batch) - 1  # 이 배치 안에서의 임시 순번 (나중에 실제 id로 치환)
                            n_claims += 1

                            prev_due = orig_due
                            for _, r in grp.iterrows():
                                this_due = safe_date(r.get("current_due_date"))
                                if this_due and prev_due and this_due != prev_due:
                                    history_batch.append({
                                        "claim_idx": claim_idx, "old": prev_due.isoformat(), "new": this_due.isoformat(),
                                        "ddays": (this_due - prev_due).days,
                                    })
                                if this_due:
                                    prev_due = this_due

                                # 이 행 자체의 예정일 + 그 행에 적힌 비고를 그대로 체크포인트로 남긴다
                                # (달력에서 특정 날짜를 볼 때, 그날 실제로 적혀있던 비고를 그대로 보여주기 위함)
                                if this_due:
                                    row_remark_raw = r.get("remark", "")
                                    row_remark = "" if pd.isna(row_remark_raw) else str(row_remark_raw).strip()
                                    if row_remark.lower() in ("none", "nan"):
                                        row_remark = ""
                                    row_unpaid_raw = r.get("sheet_unpaid_balance", None)
                                    row_unpaid = parse_amount(row_unpaid_raw) if pd.notna(row_unpaid_raw) else None
                                    checkpoint_batch.append({
                                        "claim_idx": claim_idx, "cdate": this_due.isoformat(),
                                        "remark": row_remark, "unpaid": row_unpaid,
                                    })

                            # 실제 들어온 금액(total_paid)만큼 딱 하나의 입금 기록으로 남긴다 (개별 로그 다 넣으면 취소분까지 같이 잡힘)
                            if total_paid > 0:
                                valid_pay_dates = [pd for pd, _ in payment_events if pd]
                                pdate_final = max(valid_pay_dates) if valid_pay_dates else (cur_due or date.today())
                                payment_batch.append({"claim_idx": claim_idx, "pd": pdate_final.isoformat(), "pa": total_paid})
                                if status == "완납":
                                    delay_days = 0
                                    event_type = "입금완료(정상)"
                                    if orig_due and pdate_final > orig_due:
                                        delay_days = (pdate_final - orig_due).days
                                        event_type = "입금완료(지연)"
                                    history_batch.append({
                                        "claim_idx": claim_idx, "event_pay": True, "etype": event_type,
                                        "pd": pdate_final.isoformat(), "dd": delay_days,
                                    })

                    # ---- 여기서부터 실제 DB 왕복: claims를 200개씩 나눠서 보내고 id를 순서대로 돌려받는다 ----
                    claim_ids = []
                    CHUNK = 200
                    for chunk_start in range(0, len(claims_batch), CHUNK):
                        chunk = claims_batch[chunk_start:chunk_start + CHUNK]
                        if not chunk:
                            continue
                        value_clauses = []
                        params = {}
                        for idx, cb in enumerate(chunk):
                            value_clauses.append(
                                f"(:sn{idx}, :cn{idx}, :mg{idx}, :ctype{idx}, :cdate{idx}, :odue{idx}, :cdue{idx}, :amt{idx}, :status{idx}, :remark{idx})"
                            )
                            for k, v in cb.items():
                                params[f"{k}{idx}"] = v
                        sql = f"""
                            INSERT INTO claims (site_name, company_name, manager, claim_type, claim_date, original_due_date, current_due_date, claim_amount, status, last_remark)
                            VALUES {",".join(value_clauses)}
                            RETURNING id
                        """
                        res = conn.execute(text(sql), params)
                        claim_ids.extend(row[0] for row in res.fetchall())
                        conn.commit()

                    # 자식 테이블들은 claim_idx를 실제 claim_id로 치환한 다음, 테이블당 딱 한 번씩만 왕복해서 넣는다
                    def exec_chunked(sql_text, rows, chunk_size=500):
                        for start in range(0, len(rows), chunk_size):
                            chunk = rows[start:start + chunk_size]
                            if chunk:
                                conn.execute(text(sql_text), chunk)
                                conn.commit()

                    history_rows = []
                    for h in history_batch:
                        cid = claim_ids[h["claim_idx"]]
                        if h.get("event_pay"):
                            history_rows.append({
                                "cid": cid, "etype": h["etype"], "pd": h["pd"], "dd": h["dd"],
                            })
                        else:
                            history_rows.append({
                                "cid": cid, "old": h["old"], "new": h["new"], "ddays": h["ddays"],
                            })
                    if history_rows:
                        auto_rows = [{"cid": claim_ids[h["claim_idx"]], "old": h["old"], "new": h["new"], "ddays": h["ddays"]}
                                     for h in history_batch if not h.get("event_pay")]
                        pay_rows = [{"cid": claim_ids[h["claim_idx"]], "etype": h["etype"], "pd": h["pd"], "dd": h["dd"]}
                                    for h in history_batch if h.get("event_pay")]
                        exec_chunked("""
                            INSERT INTO claim_delay_history (claim_id, event_type, old_due_date, new_due_date, delay_days, reason)
                            VALUES (:cid, '자동지연', :old, :new, :ddays, '엑셀 이력 가져오기')
                        """, auto_rows)
                        exec_chunked("""
                            INSERT INTO claim_delay_history (claim_id, event_type, payment_date, delay_days, reason)
                            VALUES (:cid, :etype, :pd, :dd, '엑셀 이력 가져오기')
                        """, pay_rows)

                    exec_chunked("""
                        INSERT INTO claim_checkpoints (claim_id, checkpoint_date, remark, unpaid_balance)
                        VALUES (:cid, :cdate, :remark, :unpaid)
                    """, [{"cid": claim_ids[c["claim_idx"]], "cdate": c["cdate"], "remark": c["remark"], "unpaid": c["unpaid"]} for c in checkpoint_batch])

                    exec_chunked(
                        "INSERT INTO payments (claim_id, payment_date, payment_amount) VALUES (:cid,:pd,:pa)",
                        [{"cid": claim_ids[p["claim_idx"]], "pd": p["pd"], "pa": p["pa"]} for p in payment_batch]
                    )

                    conn.commit()
                clear_data_cache()
                st.success(f"✅ 완료! 청구 {n_claims}건 반영 (기존 데이터는 전부 새 데이터로 갈음됨)")
                st.rerun()
            except Exception as e:
                st.error(f"❌ 처리 오류: {e}")

        st.divider()

        # ---------------- 현장별 미수관리 업로드 ----------------
        st.markdown("### 📂 현장별 미수관리 업로드 (현장별 미수현황·완불현장·계약현황)")
        recv_file = st.file_uploader("현장별 미수관리 엑셀(.xlsx) 업로드", type=["xlsx"], key="recv_upload")

        if recv_file is not None and st.button("🚀 현장별 미수관리 데이터로 전체 갱신", use_container_width=True):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(recv_file, data_only=True)
                if "미수내역" not in wb.sheetnames:
                    st.error("'미수내역' 시트를 찾지 못했습니다.")
                    st.stop()
                ws = wb["미수내역"]

                def hdr_row_idx():
                    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
                        if any(v == "공사명" for v in row):
                            return i + 1
                    return 4

                start_row = hdr_row_idx() + 1
                rows = list(ws.iter_rows(min_row=start_row, values_only=True))

                def to_date_s(v):
                    d = safe_date(v)
                    return d.isoformat() if d else None

                def is_blank(r):
                    return all(v is None for v in r)

                with engine.connect() as conn:
                    for t in ["site_receivable_details", "site_receivables", "contract_status_raw"]:
                        conn.execute(text(f"DELETE FROM {t};"))
                    conn.commit()

                    n_status_rows = 0
                    if "계약현황" in wb.sheetnames:
                        ws_cs = wb["계약현황"]
                        KEEP_LABELS = ("총 계약금", "서울", "대구", "대리점", "해외",
                                       "총 현장수", "서울현장수", "대구현장수", "대리점현장수", "해외현장수")

                        def to_f(v):
                            return float(v) if isinstance(v, (int, float)) else None

                        status_batch = []
                        for row in ws_cs.iter_rows(min_row=1, values_only=True):
                            if len(row) < 16:
                                continue
                            year_v, label_v = row[1], row[2]
                            if not isinstance(year_v, (int, float)) or label_v not in KEEP_LABELS:
                                continue
                            months = row[3:15]
                            total_v = row[15]
                            status_batch.append({
                                "yr": int(year_v), "lb": label_v,
                                **{f"m{k+1}": to_f(months[k]) for k in range(12)},
                                "tot": to_f(total_v),
                            })
                        if status_batch:
                            conn.execute(text("""
                                INSERT INTO contract_status_raw
                                (year, label, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12, total)
                                VALUES (:yr,:lb,:m1,:m2,:m3,:m4,:m5,:m6,:m7,:m8,:m9,:m10,:m11,:m12,:tot)
                            """), status_batch)
                        n_status_rows = len(status_batch)
                        conn.commit()

                    n_sites = 0
                    i = 0
                    sites_batch = []
                    details_batch_raw = []  # site_idx(임시 순번) 기준, 나중에 진짜 site_id로 치환
                    while i < len(rows):
                        r = rows[i]
                        if r[7]:  # 공사명
                            n_sites += 1
                            c_flag = r[2]
                            is_active = 1
                            status_label = "활성"
                            if isinstance(c_flag, str) and c_flag.strip() != "":
                                is_active = 0
                                status_label = c_flag.strip()
                            else:
                                # 회색 음영(테마 색상) 처리된 행 = 관리 안 하는 현장
                                fill = ws.cell(row=start_row + i, column=8).fill.fgColor
                                if getattr(fill, "type", None) == "theme":
                                    is_active = 0
                                    status_label = "미관리(회색표시)"

                            contract_date = to_date_s(r[11])
                            ym = to_date_s(r[10])
                            if not contract_date:
                                contract_date = ym

                            branch_v = r[4] or ""
                            division_v = r[5] or ""
                            if branch_v in ("해외", "대리점"):
                                division_v = branch_v  # 지사가 해외/대리점이면 구분(F열)보다 우선

                            sites_batch.append({
                                "no": int(r[6]) if isinstance(r[6], (int, float)) else None,
                                "div": division_v,
                                "sn": r[7], "cn": r[8] or "", "mg": r[24] or "", "br": branch_v,
                                "cc": r[3] or "", "cd": contract_date, "sd": to_date_s(r[12]), "ed": to_date_s(r[13]),
                                "ym": ym,
                                "ca": to_won_from_thousands(r[16]), "cha": to_won_from_thousands(r[15]),
                                "tp": to_won_from_thousands(r[17]), "ub": to_won_from_thousands(r[18]),
                                "pr": float(r[19]) if isinstance(r[19], (int, float)) else 0,
                                "ipr": float(r[20]) if isinstance(r[20], (int, float)) else 0,
                                "iir": float(r[21]) if isinstance(r[21], (int, float)) else 0,
                                "ia": is_active, "sl": status_label,
                            })
                            site_idx = len(sites_batch) - 1

                            j = i + 1
                            while j < len(rows) and not is_blank(rows[j]) and rows[j][7] is None:
                                d = rows[j]
                                # 변경계약: col11=날짜, col12=문서종류, col15=금액
                                if isinstance(d[11], datetime) and isinstance(d[15], (int, float)) and d[15]:
                                    details_batch_raw.append({"site_idx": site_idx, "dtype": "변경계약", "dt": to_date_s(d[11]),
                                                               "amt": to_won_from_thousands(d[15]), "note": d[12] or ""})
                                # 계산서: col28=발행일, col29=발행액
                                if isinstance(d[28], datetime) and isinstance(d[29], (int, float)) and d[29]:
                                    details_batch_raw.append({"site_idx": site_idx, "dtype": "계산서", "dt": to_date_s(d[28]),
                                                               "amt": to_won_from_thousands(d[29]), "note": ""})
                                # 입금: col31=입금일, col32=입금액
                                if isinstance(d[31], datetime) and isinstance(d[32], (int, float)) and d[32]:
                                    details_batch_raw.append({"site_idx": site_idx, "dtype": "입금", "dt": to_date_s(d[31]),
                                                               "amt": to_won_from_thousands(d[32]), "note": ""})
                                j += 1
                            i = j
                        else:
                            i += 1

                    # ---- 여기서부터 실제 DB 왕복: 현장을 200개씩 묶어서 나눠 보내고 id를 순서대로 받아온다.
                    #      (한 번에 774개를 다 넣으려다 문장이 너무 커져서 60초 제한에 걸려 실패했었음) ----
                    site_ids = []
                    CHUNK = 200
                    for chunk_start in range(0, len(sites_batch), CHUNK):
                        chunk = sites_batch[chunk_start:chunk_start + CHUNK]
                        if not chunk:
                            continue
                        value_clauses = []
                        params = {}
                        for idx, sb in enumerate(chunk):
                            value_clauses.append(
                                f"(:no{idx},:div{idx},:sn{idx},:cn{idx},:mg{idx},:br{idx},:cc{idx},:cd{idx},:sd{idx},:ed{idx},"
                                f":ym{idx},:ca{idx},:cha{idx},:tp{idx},:ub{idx},:pr{idx},:ipr{idx},:iir{idx},:ia{idx},:sl{idx})"
                            )
                            for k, v in sb.items():
                                params[f"{k}{idx}"] = v
                        sql = f"""
                            INSERT INTO site_receivables
                            (no_number, division, site_name, company_name, manager, branch, contract_code, contract_date, start_date,
                             completion_date, contract_yearmonth, contract_amount, change_amount, total_paid,
                             unpaid_balance, progress_rate, invoice_progress_rate, invoice_issue_rate, is_active, status_label)
                            VALUES {",".join(value_clauses)}
                            RETURNING id
                        """
                        res = conn.execute(text(sql), params)
                        site_ids.extend(row[0] for row in res.fetchall())
                        conn.commit()

                    details_batch = [
                        {"sid": site_ids[d["site_idx"]], "dtype": d["dtype"], "dt": d["dt"], "amt": d["amt"], "note": d["note"]}
                        for d in details_batch_raw
                    ]

                    for start in range(0, len(details_batch), 500):
                        chunk = details_batch[start:start + 500]
                        if chunk:
                            conn.execute(text("""
                                INSERT INTO site_receivable_details (site_receivable_id, detail_type, detail_date, amount, note)
                                VALUES (:sid, :dtype, :dt, :amt, :note)
                            """), chunk)
                            conn.commit()

                    conn.commit()
                clear_data_cache()
                st.success(f"✅ 완료! 현장 {n_sites}개, 계약현황 {n_status_rows}행 반영 (기존 데이터는 전부 새 데이터로 갈음됨)")
                st.rerun()
            except Exception as e:
                st.error(f"❌ 처리 오류: {e}")